from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def _money(
    value,
):

    return (
        f"INR {Decimal(str(value or '0')).quantize(Decimal('0.01')):,.2f}"
    )


def _summary_rows(
    summary,
):

    return [
        [
            "Orders",
            str(
                summary.get(
                    "order_count",
                    0,
                ),
            ),
        ],
        [
            "Gross sales",
            _money(
                summary.get(
                    "subtotal_gross",
                ),
            ),
        ],
        [
            "Offer savings",
            _money(
                summary.get(
                    "offer_discount_total",
                ),
            ),
        ],
        [
            "Coupon deductions",
            _money(
                summary.get(
                    "coupon_discount_total",
                ),
            ),
        ],
        [
            "Total discounts",
            _money(
                summary.get(
                    "total_discount",
                ),
            ),
        ],
        [
            "Tax",
            _money(
                summary.get(
                    "tax_total",
                ),
            ),
        ],
        [
            "Shipping",
            _money(
                summary.get(
                    "shipping_total",
                ),
            ),
        ],
        [
            "Net sales",
            _money(
                summary.get(
                    "grand_total",
                ),
            ),
        ],
    ]


def build_sales_report_pdf(
    report_data,
):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(
            A4,
        ),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="FurniCart Sales Report",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Heading1"],
        fontSize=15,
        spaceAfter=8,
        textColor=colors.HexColor(
            "#2c241c",
        ),
    )

    muted = ParagraphStyle(
        name="ReportMuted",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor(
            "#6b635c",
        ),
    )

    story = [
        Paragraph(
            "FurniCart Sales Report",
            title_style,
        ),
        Paragraph(
            escape(
                f"Period: {report_data['period']} "
                f"({report_data['date_from']} to {report_data['date_to']})",
            ),
            muted,
        ),
        Paragraph(
            escape(
                f"Generated {timezone.localtime().strftime('%Y-%m-%d %H:%M')}",
            ),
            muted,
        ),
        Spacer(
            1,
            0.35 * cm,
        ),
    ]

    summary_table = Table(
        [
            [
                "Metric",
                "Value",
            ],
            *_summary_rows(
                report_data[
                    "summary"
                ],
            ),
        ],
        colWidths=[
            8 * cm,
            6 * cm,
        ],
    )

    summary_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        0,
                    ),
                    colors.HexColor(
                        "#4b2d1d",
                    ),
                ),
                (
                    "TEXTCOLOR",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        0,
                    ),
                    colors.white,
                ),
                (
                    "GRID",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    0.25,
                    colors.HexColor(
                        "#e7e2dc",
                    ),
                ),
                (
                    "FONTNAME",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        0,
                    ),
                    "Helvetica-Bold",
                ),
                (
                    "FONTSIZE",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    8,
                ),
            ],
        ),
    )

    story.append(
        summary_table,
    )

    story.append(
        Spacer(
            1,
            0.4 * cm,
        ),
    )

    breakdown = report_data.get(
        "breakdown",
        [],
    )

    if breakdown:

        story.append(
            Paragraph(
                "Breakdown",
                title_style,
            ),
        )

        bd_rows = [
            [
                "Period",
                "Orders",
                "Gross",
                "Offers",
                "Coupons",
                "Net sales",
            ],
        ]

        for row in breakdown:

            bd_rows.append(
                [
                    row.get(
                        "label",
                        "",
                    ),
                    str(
                        row.get(
                            "order_count",
                            0,
                        ),
                    ),
                    _money(
                        row.get(
                            "subtotal_gross",
                        ),
                    ),
                    _money(
                        row.get(
                            "offer_discount_total",
                        ),
                    ),
                    _money(
                        row.get(
                            "coupon_discount_total",
                        ),
                    ),
                    _money(
                        row.get(
                            "grand_total",
                        ),
                    ),
                ],
            )

        bd_table = Table(
            bd_rows,
            repeatRows=1,
        )

        bd_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            0,
                        ),
                        colors.HexColor(
                            "#4b2d1d",
                        ),
                    ),
                    (
                        "TEXTCOLOR",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            0,
                        ),
                        colors.white,
                    ),
                    (
                        "GRID",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            -1,
                        ),
                        0.25,
                        colors.HexColor(
                            "#e7e2dc",
                        ),
                    ),
                    (
                        "FONTSIZE",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            -1,
                        ),
                        7,
                    ),
                ],
            ),
        )

        story.append(
            bd_table,
        )

        story.append(
            Spacer(
                1,
                0.4 * cm,
            ),
        )

    orders = report_data.get(
        "orders",
        [],
    )

    if orders:

        story.append(
            Paragraph(
                "Orders",
                title_style,
            ),
        )

        order_rows = [
            [
                "Order",
                "Date",
                "Customer",
                "Payment",
                "Offers",
                "Coupon",
                "Total",
            ],
        ]

        for order in orders:

            order_rows.append(
                [
                    order.get(
                        "order_number",
                        "",
                    ),
                    order.get(
                        "placed_at",
                        "",
                    )[
                        :16
                    ],
                    order.get(
                        "customer_email",
                        "",
                    ),
                    order.get(
                        "payment_method",
                        "",
                    ),
                    _money(
                        order.get(
                            "offer_discount_total",
                        ),
                    ),
                    _money(
                        order.get(
                            "coupon_discount_total",
                        ),
                    ),
                    _money(
                        order.get(
                            "grand_total",
                        ),
                    ),
                ],
            )

        orders_table = Table(
            order_rows,
            repeatRows=1,
        )

        orders_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            0,
                        ),
                        colors.HexColor(
                            "#4b2d1d",
                        ),
                    ),
                    (
                        "TEXTCOLOR",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            0,
                        ),
                        colors.white,
                    ),
                    (
                        "GRID",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            -1,
                        ),
                        0.25,
                        colors.HexColor(
                            "#e7e2dc",
                        ),
                    ),
                    (
                        "FONTSIZE",
                        (
                            0,
                            0,
                        ),
                        (
                            -1,
                            -1,
                        ),
                        7,
                    ),
                ],
            ),
        )

        story.append(
            orders_table,
        )

    doc.build(
        story,
    )

    return buffer.getvalue()


def build_sales_report_excel(
    report_data,
):

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(
        [
            "FurniCart Sales Report",
        ],
    )

    ws_summary.append(
        [
            "Period",
            report_data[
                "period"
            ],
        ],
    )

    ws_summary.append(
        [
            "From",
            report_data[
                "date_from"
            ],
        ],
    )

    ws_summary.append(
        [
            "To",
            report_data[
                "date_to"
            ],
        ],
    )

    ws_summary.append(
        [],
    )

    ws_summary.append(
        [
            "Metric",
            "Value",
        ],
    )

    for label, value in _summary_rows(
        report_data[
            "summary"
        ],
    ):

        ws_summary.append(
            [
                label,
                value,
            ],
        )

    ws_summary[
        "A1"
    ].font = Font(
        bold=True,
        size=14,
    )

    ws_breakdown = wb.create_sheet(
        "Breakdown",
    )

    ws_breakdown.append(
        [
            "Period",
            "Orders",
            "Gross sales",
            "Offer savings",
            "Coupon deductions",
            "Total discounts",
            "Tax",
            "Shipping",
            "Net sales",
        ],
    )

    for row in report_data.get(
        "breakdown",
        [],
    ):

        ws_breakdown.append(
            [
                row.get(
                    "label",
                ),
                row.get(
                    "order_count",
                ),
                row.get(
                    "subtotal_gross",
                ),
                row.get(
                    "offer_discount_total",
                ),
                row.get(
                    "coupon_discount_total",
                ),
                row.get(
                    "total_discount",
                ),
                row.get(
                    "tax_total",
                ),
                row.get(
                    "shipping_total",
                ),
                row.get(
                    "grand_total",
                ),
            ],
        )

    ws_orders = wb.create_sheet(
        "Orders",
    )

    ws_orders.append(
        [
            "Order",
            "Placed at",
            "Customer",
            "Status",
            "Payment",
            "Gross",
            "Offers",
            "Subtotal",
            "Coupon",
            "Coupon code",
            "Tax",
            "Shipping",
            "Grand total",
        ],
    )

    for order in report_data.get(
        "orders",
        [],
    ):

        ws_orders.append(
            [
                order.get(
                    "order_number",
                ),
                order.get(
                    "placed_at",
                ),
                order.get(
                    "customer_email",
                ),
                order.get(
                    "status",
                ),
                order.get(
                    "payment_method",
                ),
                order.get(
                    "subtotal_gross",
                ),
                order.get(
                    "offer_discount_total",
                ),
                order.get(
                    "subtotal",
                ),
                order.get(
                    "coupon_discount_total",
                ),
                order.get(
                    "coupon_code",
                ),
                order.get(
                    "tax_total",
                ),
                order.get(
                    "shipping_total",
                ),
                order.get(
                    "grand_total",
                ),
            ],
        )

    buffer = BytesIO()

    wb.save(
        buffer,
    )

    return buffer.getvalue()


def build_ledger_excel(
    report_data,
):

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "Ledger"

    ws.append(
        [
            "FurniCart Sales Ledger",
        ],
    )

    ws.append(
        [
            "From",
            report_data[
                "date_from"
            ],
            "To",
            report_data[
                "date_to"
            ],
        ],
    )

    ws.append(
        [],
    )

    ws.append(
        [
            "Date",
            "Order",
            "Customer",
            "Payment method",
            "Payment status",
            "Gross",
            "Offer discount",
            "Coupon discount",
            "Tax",
            "Shipping",
            "Amount (IN)",
        ],
    )

    for order in report_data.get(
        "orders",
        [],
    ):

        ws.append(
            [
                order.get(
                    "placed_at",
                ),
                order.get(
                    "order_number",
                ),
                order.get(
                    "customer_email",
                ),
                order.get(
                    "payment_method",
                ),
                order.get(
                    "payment_status",
                ),
                order.get(
                    "subtotal_gross",
                ),
                order.get(
                    "offer_discount_total",
                ),
                order.get(
                    "coupon_discount_total",
                ),
                order.get(
                    "tax_total",
                ),
                order.get(
                    "shipping_total",
                ),
                order.get(
                    "grand_total",
                ),
            ],
        )

    ws[
        "A1"
    ].font = Font(
        bold=True,
        size=14,
    )

    buffer = BytesIO()

    wb.save(
        buffer,
    )

    return buffer.getvalue()


def build_ledger_pdf(
    report_data,
):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(
            A4,
        ),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="FurniCart Sales Ledger",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="LedgerTitle",
        parent=styles["Heading1"],
        fontSize=14,
        spaceAfter=6,
    )

    story = [
        Paragraph(
            "FurniCart Sales Ledger",
            title_style,
        ),
        Paragraph(
            escape(
                f"{report_data['date_from']} to {report_data['date_to']}",
            ),
        ),
        Spacer(
            1,
            0.3 * cm,
        ),
    ]

    rows = [
        [
            "Date",
            "Order",
            "Customer",
            "Payment",
            "Gross",
            "Discounts",
            "Amount",
        ],
    ]

    for order in report_data.get(
        "orders",
        [],
    ):

        discounts = Decimal(
            order.get(
                "total_discount",
                "0",
            ),
        )

        rows.append(
            [
                order.get(
                    "placed_at",
                    "",
                )[
                    :10
                ],
                order.get(
                    "order_number",
                    "",
                ),
                order.get(
                    "customer_email",
                    "",
                ),
                order.get(
                    "payment_method",
                    "",
                ),
                _money(
                    order.get(
                        "subtotal_gross",
                    ),
                ),
                _money(
                    discounts,
                ),
                _money(
                    order.get(
                        "grand_total",
                    ),
                ),
            ],
        )

    table = Table(
        rows,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        0,
                    ),
                    colors.HexColor(
                        "#4b2d1d",
                    ),
                ),
                (
                    "TEXTCOLOR",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        0,
                    ),
                    colors.white,
                ),
                (
                    "GRID",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    0.25,
                    colors.HexColor(
                        "#e7e2dc",
                    ),
                ),
                (
                    "FONTSIZE",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    7,
                ),
            ],
        ),
    )

    story.append(
        table,
    )

    doc.build(
        story,
    )

    return buffer.getvalue()
